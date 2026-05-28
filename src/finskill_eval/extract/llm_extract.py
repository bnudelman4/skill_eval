"""LLM extraction of a skill artifact -> Ledger (Option C).

Replaces the brittle layout-guessing xlsx parser on the critical path. A cheap
LLM reads the spreadsheet (rendered to text) and returns a fixed-schema JSON
array of cells; trivial deterministic code loads it and `build_ledger` runs the
SAME normalization/wiring as the xlsx parser. The skill under test is untouched;
the extractor is part of the grader. Every extracted number is still checked by
Python (recompute + compare vs gold), so an extraction slip cannot pass silently.

No per-layout code: a new spreadsheet shape is handled by the LLM, not by new
regexes. The extractor prompt + schema are stable and iterated OFFLINE on cached
artifacts (zero live usage).
"""

from __future__ import annotations

import json
import os
import re
import subprocess
from pathlib import Path
from typing import Callable, Optional

from openpyxl import load_workbook

from finskill_eval.ledger import Ledger
from finskill_eval.parse_xlsx import build_ledger

LLM = Callable[[str], str]

_SCHEMA_PROMPT = """You are a data-extraction tool. Below is a financial
spreadsheet rendered as text (one or more sheets). Extract EVERY numeric metric
into a JSON array. For each cell output an object:

  {"metric": "<canonical snake_case name, e.g. revenue, gross_profit,
              gross_margin, operating_cash_flow, share_repurchases>",
   "period": "<FY2024 or 'Q4 2024'>",
   "value": <number in ABSOLUTE units — expand millions/thousands to the full
             number; percentages as the percent value e.g. 46.21>,
   "unit": "<one of: USD, percent, shares, ratio>",
   "kind": "<direct if sourced, derived if it is a ratio/margin/growth/multiple>",
   "fmp_endpoint": "<the FMP endpoint name from the catalog below where the
                     CANONICAL value for this metric would be found; null if
                     no FMP endpoint carries this metric>",
   "fmp_field":    "<the exact field name within fmp_endpoint that holds this
                     metric (case-sensitive); null if not in FMP>"}

Rules:
- ABSOLUTE values only: if the sheet shows revenue as 391035 ($ millions),
  output 391035000000.
- Use canonical snake_case metric names; do not invent units.
- For fmp_endpoint/fmp_field: ONLY use a (endpoint, field) pair that appears
  verbatim in the FMP CATALOG below. If a metric is computed/derived in the
  skill but FMP publishes the same metric directly, use FMP's pre-computed
  field rather than the inputs (e.g. gross_margin -> ratios.grossProfitMargin).
  If the metric is bank/segment-specific or doesn't have an FMP equivalent,
  set both fields to null.
- Use the SKILL CONTEXT to disambiguate: skill-specific names like
  "NII as % of Revenue" or "Diluted Shares (M)" or "D&A" map to the standard
  FMP fields (interestIncome / revenue, weightedAverageShsOutDil,
  depreciationAndAmortization) — propose the mapping.
- Output ONLY the JSON array, no prose, no code fences.

SKILL CONTEXT (what the skill is computing):
__SKILL_MD__

FMP CATALOG (what FMP actually exposes — pick fields ONLY from here):
__FMP_CATALOG__

SPREADSHEET:
__SHEET_TEXT__
"""


def xlsx_to_text(path: str | Path) -> str:
    """Render every sheet to a compact tab-separated text grid for the LLM."""
    wb = load_workbook(Path(path), data_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        out.append(f"=== SHEET: {ws.title} ===")
        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                out.append("\t".join("" if c is None else str(c) for c in row))
    return "\n".join(out)


def parse_extractor_json(raw: str) -> list[dict]:
    """Tolerant JSON load: strip code fences / prose, find the JSON array."""
    s = raw.strip()
    if "```" in s:
        s = re.sub(r"```(?:json)?", "", s).strip()
    try:
        obj = json.loads(s)
    except json.JSONDecodeError:
        m = re.search(r"\[.*\]", s, re.DOTALL)
        if not m:
            return []
        try:
            obj = json.loads(m.group(0))
        except json.JSONDecodeError:
            return []
    return obj if isinstance(obj, list) else []


def _records_to_raw_cells(records: list[dict]) -> list[dict]:
    raw_cells = []
    for r in records:
        if "metric" not in r or "value" not in r:
            continue
        raw_cells.append({
            "label": str(r.get("metric", "")),
            "period_raw": str(r.get("period", "")),
            "value_raw": r.get("value"),
            "unit": str(r.get("unit", "")),
            # Per-cell FMP mapping the extractor proposed (SKILL.md +
            # catalog-aware). None when the LLM declined to map.
            "fmp_endpoint": r.get("fmp_endpoint") or None,
            "fmp_field": r.get("fmp_field") or None,
        })
    return raw_cells


def _default_llm(prompt: str) -> str:
    """Cheap Haiku call via subscription auth (no metered $, no 30k-TPM cap)."""
    env = {k: v for k, v in os.environ.items() if k != "ANTHROPIC_API_KEY"}
    proc = subprocess.run(
        ["claude", "-p", prompt, "--output-format", "json",
         "--model", "claude-haiku-4-5-20251001", "--max-turns", "1"],
        capture_output=True, text=True, env=env, timeout=180,
    )
    try:
        return json.loads(proc.stdout).get("result", "")
    except (json.JSONDecodeError, AttributeError):
        return proc.stdout


def extract_ledger(
    path: str | Path,
    *,
    skill: str,
    ticker: str,
    llm: Optional[LLM] = None,
    skill_md: Optional[str] = None,
    fmp_catalog: Optional[dict[str, list[str]]] = None,
) -> Ledger:
    """Extract a Ledger from an xlsx artifact via LLM.

    When `skill_md` (the skill's instruction body) and `fmp_catalog` (FMP's
    full field universe) are provided, the extractor ALSO proposes a per-cell
    FMP mapping (endpoint + field) inline. The verifier then uses those
    mappings directly — no LABEL_MAP lookups, no separate resolver pass.

    When either is omitted, falls back to the original prompt and downstream
    LABEL_MAP / resolver chain.
    """
    llm = llm or _default_llm
    text = xlsx_to_text(path)
    prompt = _SCHEMA_PROMPT.replace("__SHEET_TEXT__", text)
    prompt = prompt.replace(
        "__SKILL_MD__",
        skill_md if skill_md else "(no SKILL.md provided)",
    )
    if fmp_catalog:
        flat = "\n".join(
            f"{ep}: {', '.join(sorted(fields))}"
            for ep, fields in sorted(fmp_catalog.items())
        )
    else:
        flat = "(no FMP catalog provided)"
    prompt = prompt.replace("__FMP_CATALOG__", flat)
    raw = llm(prompt)
    records = parse_extractor_json(raw)
    return build_ledger(_records_to_raw_cells(records), skill=skill, ticker=ticker)
