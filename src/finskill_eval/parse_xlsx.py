"""xlsx artifact -> Ledger (M2.4).

Walks the sheet, forward-filling label/period/unit columns so merged cells and
sparse layouts still pair each value with its nearest descriptive label.
Normalizes values, applies unit-based scaling to canonical base units, and
classifies direct vs. derived. Derived cells are wired to their input cells via
the metric-definition registry so recompute() can run later.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from finskill_eval.ledger import Cell, Ledger
from finskill_eval.normalize import (
    normalize_label,
    normalize_number,
    normalize_period,
    period_key,
)
from finskill_eval.recompute import METRIC_DEFS

# Unit string -> multiplier into canonical base units (absolute $, share count).
# Percent and multiples are already dimensionless.
UNIT_SCALE: dict[str, float] = {
    "$mm": 1_000_000.0,
    "$m": 1_000_000.0,
    "mm": 1_000_000.0,
    "$bn": 1_000_000_000.0,
    "thousands": 1_000.0,
    "$k": 1_000.0,
    "%": 1.0,
    "x": 1.0,
    # canonical absolute units emitted by the LLM extractor (already absolute)
    "usd": 1.0,
    "USD": 1.0,
    "percent": 1.0,
    "shares": 1.0,
    "ratio": 1.0,
    "": 1.0,
}

# Editable keyword fallback for classifying derived cells when a metric is not
# (yet) in METRIC_DEFS. METRIC_DEFS membership is authoritative.
_DERIVED_KEYWORDS = re.compile(r"margin|yield|growth|ratio|multiple|ev_ebitda|pe_")
_DERIVED_UNITS = {"x"}


def _classify_kind(canonical_label: str, unit: str, is_numeric: bool) -> str:
    if not is_numeric:
        return "direct"
    if canonical_label in METRIC_DEFS:
        return "derived"
    if _DERIVED_KEYWORDS.search(canonical_label) or unit in _DERIVED_UNITS:
        return "derived"
    return "direct"


def _faith_cell_type(kind: str, formula: str | None, n_inputs: int) -> str:
    """FAITH 4-type taxonomy: direct lookup / comparative / bivariate / multivariate."""
    if kind == "direct":
        return "direct_lookup"
    if formula == "growth":
        return "comparative"  # same metric across periods
    if n_inputs >= 3:
        return "multivariate"
    return "bivariate"  # two distinct metrics


def _safe_period(raw: object):
    """normalize_period but tolerant: unparseable -> None instead of raising,
    so one odd cell never crashes a whole multi-sheet workbook."""
    try:
        return normalize_period(raw)
    except ValueError:
        return None


_PCT_RATIO_LABEL = re.compile(r"%|percent|margin|yield|ratio|growth|change|rate")
_MILLIONS = re.compile(r"million|\$mm|\(\$m\)|\$ m\b", re.IGNORECASE)


def _sheet_unit(ws) -> str:
    """Infer a sheet-wide value unit from its title / header text."""
    blob = " ".join(
        str(c) for row in ws.iter_rows(min_row=1, max_row=3, values_only=True)
        for c in row if c is not None
    )
    return "$mm" if _MILLIONS.search(blob) else ""


def _find_matrix_header(ws):
    """Return (row_idx, {col_idx: period_raw}) if a header row has >=2 columns
    that parse as periods; else None. Scans the first several rows."""
    rows = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
    for i, row in enumerate(rows):
        periods = {
            j: str(c)
            for j, c in enumerate(row)
            if j >= 1 and c is not None and _safe_period(c) is not None
        }
        if len(periods) >= 2:
            return i, periods
    return None


def _extract_matrix(ws, header_row: int, periods: dict, sheet_unit: str) -> list[dict]:
    out: list[dict] = []
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[header_row + 1:]:
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue
        label = str(row[0])
        unit = "" if _PCT_RATIO_LABEL.search(label.lower()) else sheet_unit
        for col, period_raw in periods.items():
            if col >= len(row):
                continue
            val = row[col]
            if val is None or str(val).strip() == "":
                continue
            out.append({"label": label, "period_raw": period_raw,
                        "value_raw": val, "unit": unit})
    return out


def _extract_row_oriented(ws) -> list[dict]:
    label_fill = period_fill = unit_fill = ""
    out: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        cols = list(row) + [None] * (4 - len(row))
        metric, period_raw, value_raw, unit_raw = cols[:4]
        if value_raw is None or str(value_raw).strip() == "":
            continue
        if str(value_raw).strip().lower() == "value":
            continue
        if metric not in (None, ""):
            label_fill = str(metric)
        if period_raw not in (None, ""):
            period_fill = str(period_raw)
        if unit_raw not in (None, ""):
            unit_fill = str(unit_raw)
        out.append({"label": label_fill, "period_raw": period_fill,
                    "value_raw": value_raw, "unit": unit_fill})
    return out


def parse(path: str | Path, *, skill: str, ticker: str) -> Ledger:
    wb = load_workbook(Path(path), data_only=True)

    raw_cells: list[dict] = []
    for ws in wb.worksheets:
        matrix = _find_matrix_header(ws)
        if matrix is not None:
            header_row, periods = matrix
            raw_cells += _extract_matrix(ws, header_row, periods, _sheet_unit(ws))
        else:
            raw_cells += _extract_row_oriented(ws)

    return build_ledger(raw_cells, skill=skill, ticker=ticker)


def build_ledger(raw_cells: list[dict], *, skill: str, ticker: str) -> Ledger:
    """Normalize + classify + wire a list of {label, period_raw, value_raw, unit}
    records into a Ledger. Shared by the xlsx parser and the LLM extractor so
    both feed the identical downstream pipeline."""
    cells: list[Cell] = []
    by_label_period: dict[tuple[str, str | None], str] = {}
    seen_ids: set[str] = set()
    for rc in raw_cells:
        canonical = normalize_label(rc["label"])
        period = _safe_period(rc["period_raw"])
        pkey = period_key(period)
        num = normalize_number(rc["value_raw"])
        unit = rc["unit"]

        if num is None:
            value: float | str = str(rc["value_raw"]).strip()  # non-numeric (exact)
            is_numeric = False
        else:
            value = num * UNIT_SCALE.get(unit, 1.0)
            is_numeric = True

        kind = _classify_kind(canonical, unit, is_numeric)
        cell_id = f"{canonical}__{pkey}" if pkey else canonical
        if cell_id in seen_ids:
            continue  # same metric repeated across sheets -> keep first
        seen_ids.add(cell_id)
        cells.append(
            Cell(
                cell_id=cell_id,
                label=rc["label"],
                canonical_label=canonical,
                period=period,
                raw_value=str(rc["value_raw"]),
                value=value,
                unit=unit,
                kind=kind,
                cell_type="direct_lookup" if kind == "direct" else None,
            )
        )
        by_label_period[(canonical, pkey)] = cell_id

    # wire derived cells to their input cells (same period) via METRIC_DEFS
    wired: list[Cell] = []
    for c in cells:
        if c.kind == "derived" and c.canonical_label in METRIC_DEFS:
            formula, input_labels = METRIC_DEFS[c.canonical_label]
            pkey2 = c.cell_id.split("__", 1)[1] if "__" in c.cell_id else None
            input_ids = tuple(
                by_label_period[(lbl, pkey2)]
                for lbl in input_labels
                if (lbl, pkey2) in by_label_period
            )
            cell_type = _faith_cell_type("derived", formula, len(input_ids))
            wired.append(
                c.model_copy(
                    update={
                        "formula": formula,
                        "inputs": input_ids,
                        "cell_type": cell_type,
                    }
                )
            )
        else:
            wired.append(c)

    return Ledger(skill=skill, ticker=ticker, cells=wired)
